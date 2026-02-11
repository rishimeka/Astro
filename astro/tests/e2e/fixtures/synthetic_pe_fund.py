"""Enhanced synthetic PE fund model for E2E testing.

This fixture creates a realistic fund model that exercises ALL pattern types
from Section 10.3 of the POC spec:

1. Cross-sheet reference - Fees reference rates from Assumptions sheet
2. Time-series growth - Revenue grows at constant rate per period
3. Conditional formula - Distributions only occur when NAV > hurdle
4. Cumulative row - Cumulative distributions running total
5. Clear input vs. calculated separation - Clearly marked input and formula rows

Structure:
- Assumptions sheet: Input parameters (rates, hurdle, initial values)
- Cash Flows sheet: Capital calls, distributions, revenue, fees, net
- NAV sheet: NAV calculation with conditional distributions

This model is designed to require multiple interview iterations to fully
capture all relationships, exercising the EvalStar loop.
"""

import os
import tempfile
from collections.abc import Generator
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_synthetic_pe_fund(output_path: str | None = None) -> str:
    """Create a synthetic PE fund model .xlsx file for E2E testing.

    Args:
        output_path: Optional path to save the file. If None, uses a temp file.

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Styling for clarity
    header_font = Font(bold=True)
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    calc_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue

    # =========================================================================
    # ASSUMPTIONS SHEET - All input parameters
    # =========================================================================
    assumptions = wb.create_sheet("Assumptions")

    # Headers
    assumptions["A1"] = "Parameter"
    assumptions["B1"] = "Value"
    assumptions["C1"] = "Description"
    for cell in [assumptions["A1"], assumptions["B1"], assumptions["C1"]]:
        cell.font = header_font

    # Input parameters (rows 3-10)
    params = [
        ("Growth Rate", 0.05, "Annual revenue growth rate"),
        ("Management Fee Rate", 0.02, "Annual management fee as % of committed capital"),
        ("Carry Rate", 0.20, "Carried interest rate on profits above hurdle"),
        ("Hurdle Rate", 0.08, "Preferred return hurdle for carry calculation"),
        ("Initial Revenue", 1000000, "Starting revenue in Q1"),
        ("Committed Capital", 50000000, "Total committed capital"),
        ("Investment Period", 5, "Number of quarters for capital calls"),
        ("Fund Life", 11, "Total quarters in fund life"),
    ]

    for i, (name, value, desc) in enumerate(params, start=3):
        assumptions.cell(row=i, column=1, value=name)
        cell = assumptions.cell(row=i, column=2, value=value)
        cell.fill = input_fill  # Mark as input
        assumptions.cell(row=i, column=3, value=desc)

    # =========================================================================
    # CASH FLOWS SHEET - Time series with multiple patterns
    # =========================================================================
    cash_flows = wb.create_sheet("Cash Flows")

    # Period headers (row 1)
    cash_flows["A1"] = "Metric"
    cash_flows["A1"].font = header_font
    periods = 11  # Q1-Q11
    for col in range(2, periods + 2):
        cell = cash_flows.cell(row=1, column=col, value=f"Q{col-1}")
        cell.font = header_font

    # Row structure:
    # Row 3: Capital Calls (INPUT) - calls during investment period
    # Row 4: Distributions (INPUT for base, CONDITIONAL for actual)
    # Row 5: Revenue (CALCULATED - time series growth)
    # Row 6: Management Fees (CALCULATED - cross-sheet reference)
    # Row 7: Net Operating Income (CALCULATED - arithmetic)
    # Row 8: Cumulative Distributions (CALCULATED - cumulative)
    # Row 9: Cumulative Capital Calls (CALCULATED - cumulative)

    row_labels = {
        3: ("Capital Calls", "input"),
        4: ("Distributions", "conditional"),  # Conditional based on NAV
        5: ("Revenue", "time_series"),
        6: ("Management Fees", "cross_sheet"),
        7: ("Net Operating Income", "arithmetic"),
        8: ("Cumulative Distributions", "cumulative"),
        9: ("Cumulative Capital Calls", "cumulative"),
    }

    for row, (label, pattern_type) in row_labels.items():
        cell = cash_flows.cell(row=row, column=1, value=label)
        if pattern_type == "input":
            cell.fill = input_fill
        else:
            cell.fill = calc_fill

    # Generate values
    committed_capital = 50000000
    growth_rate = 0.05
    mgmt_fee_rate = 0.02
    hurdle_rate = 0.08
    initial_revenue = 1000000
    investment_period = 5

    # Track cumulative values
    cum_capital_calls = 0.0
    cum_distributions = 0.0
    nav_values: list[float] = []

    for col in range(2, periods + 2):
        quarter = col - 1

        # Row 3: Capital Calls (INPUT)
        # Calls happen in first 5 quarters, equal installments
        if quarter <= investment_period:
            capital_call = committed_capital / investment_period
        else:
            capital_call = 0
        cash_flows.cell(row=3, column=col, value=capital_call).fill = input_fill
        cum_capital_calls += capital_call

        # Row 5: Revenue (TIME SERIES GROWTH)
        # Revenue = Initial * (1 + growth_rate)^(quarter-1)
        revenue = initial_revenue * ((1 + growth_rate) ** (quarter - 1))
        cash_flows.cell(row=5, column=col, value=revenue).fill = calc_fill

        # Row 6: Management Fees (CROSS-SHEET REFERENCE)
        # Fees = Committed Capital * Fee Rate / 4 (quarterly)
        mgmt_fee = committed_capital * mgmt_fee_rate / 4
        cash_flows.cell(row=6, column=col, value=mgmt_fee).fill = calc_fill

        # Row 7: Net Operating Income (ARITHMETIC)
        # NOI = Revenue - Management Fees
        noi = revenue - mgmt_fee
        cash_flows.cell(row=7, column=col, value=noi).fill = calc_fill

        # Calculate NAV for conditional distribution logic
        # Simplified: NAV = Cumulative Capital Calls + Cumulative NOI - Cumulative Distributions
        if col == 2:
            nav = capital_call + noi
        else:
            prev_nav = nav_values[-1] if nav_values else 0
            nav = prev_nav + capital_call + noi - cum_distributions

        # Row 4: Distributions (CONDITIONAL)
        # Distribution only if NAV > (Cumulative Calls * (1 + Hurdle Rate))
        hurdle_amount = cum_capital_calls * (1 + hurdle_rate * (quarter / 4))
        if nav > hurdle_amount and quarter > investment_period:
            # Distribute excess over hurdle
            distribution = min((nav - hurdle_amount) * 0.5, nav * 0.3)  # Max 30% of NAV
        else:
            distribution = 0
        cash_flows.cell(row=4, column=col, value=distribution).fill = calc_fill
        cum_distributions += distribution

        # Update NAV after distribution
        nav = nav - distribution
        nav_values.append(nav)

        # Row 8: Cumulative Distributions (CUMULATIVE)
        cash_flows.cell(row=8, column=col, value=cum_distributions).fill = calc_fill

        # Row 9: Cumulative Capital Calls (CUMULATIVE)
        cash_flows.cell(row=9, column=col, value=cum_capital_calls).fill = calc_fill

    # =========================================================================
    # NAV SHEET - More complex calculations with conditionals
    # =========================================================================
    nav_sheet = wb.create_sheet("NAV")

    # Headers
    nav_sheet["A1"] = "Metric"
    nav_sheet["A1"].font = header_font
    for col in range(2, periods + 2):
        cell = nav_sheet.cell(row=1, column=col, value=f"Q{col-1}")
        cell.font = header_font

    # Row structure for NAV:
    # Row 3: Beginning NAV (prior period ending NAV)
    # Row 4: + Capital Calls (cross-sheet reference)
    # Row 5: + NOI (cross-sheet reference)
    # Row 6: - Distributions (cross-sheet reference)
    # Row 7: = Ending NAV (arithmetic)
    # Row 8: Hurdle Amount (cross-sheet calculation)
    # Row 9: Above Hurdle? (conditional - TRUE/FALSE)

    nav_labels = {
        3: ("Beginning NAV", "time_series"),
        4: ("+ Capital Calls", "cross_sheet"),
        5: ("+ NOI", "cross_sheet"),
        6: ("- Distributions", "cross_sheet"),
        7: ("Ending NAV", "arithmetic"),
        8: ("Hurdle Amount", "cross_sheet"),
        9: ("Above Hurdle?", "conditional"),
    }

    for row, (label, _) in nav_labels.items():
        cell = nav_sheet.cell(row=row, column=1, value=label)
        cell.fill = calc_fill

    # Generate NAV values
    for col in range(2, periods + 2):
        quarter = col - 1

        # Beginning NAV (prior ending, or 0 for Q1)
        if col == 2:
            beginning_nav = 0
        else:
            beginning_nav = nav_sheet.cell(row=7, column=col - 1).value or 0
        nav_sheet.cell(row=3, column=col, value=beginning_nav)

        # Capital Calls (from Cash Flows)
        capital_calls = cash_flows.cell(row=3, column=col).value
        nav_sheet.cell(row=4, column=col, value=capital_calls)

        # NOI (from Cash Flows)
        noi = cash_flows.cell(row=7, column=col).value
        nav_sheet.cell(row=5, column=col, value=noi)

        # Distributions (from Cash Flows)
        distributions = cash_flows.cell(row=4, column=col).value
        nav_sheet.cell(row=6, column=col, value=distributions)

        # Ending NAV
        ending_nav = beginning_nav + capital_calls + noi - distributions
        nav_sheet.cell(row=7, column=col, value=ending_nav)

        # Hurdle Amount
        cum_calls = cash_flows.cell(row=9, column=col).value
        hurdle = cum_calls * (1 + hurdle_rate * (quarter / 4))
        nav_sheet.cell(row=8, column=col, value=hurdle)

        # Above Hurdle? (CONDITIONAL)
        above_hurdle = ending_nav > hurdle
        nav_sheet.cell(row=9, column=col, value="TRUE" if above_hurdle else "FALSE")

    # =========================================================================
    # Save the workbook
    # =========================================================================
    if output_path is None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

    wb.save(output_path)
    wb.close()

    return output_path


def create_synthetic_pe_fund_fixture() -> Generator[str, None, None]:
    """Pytest fixture version that cleans up after test."""
    path = create_synthetic_pe_fund()
    yield path
    if os.path.exists(path):
        os.unlink(path)


# Pattern documentation for interview answers
EXPECTED_PATTERNS: dict[str, Any] = {
    "Assumptions": {
        "type": "input_sheet",
        "input_cells": [
            ("B3", "Growth Rate", 0.05),
            ("B4", "Management Fee Rate", 0.02),
            ("B5", "Carry Rate", 0.20),
            ("B6", "Hurdle Rate", 0.08),
            ("B7", "Initial Revenue", 1000000),
            ("B8", "Committed Capital", 50000000),
            ("B9", "Investment Period", 5),
            ("B10", "Fund Life", 11),
        ],
    },
    "Cash Flows": {
        "type": "calculation_sheet",
        "patterns": [
            {
                "row": 3,
                "label": "Capital Calls",
                "pattern_type": "input",
                "description": "Equal installments over investment period (5 quarters)",
            },
            {
                "row": 4,
                "label": "Distributions",
                "pattern_type": "conditional",
                "description": "Distribution if NAV > Hurdle Amount AND past investment period",
                "formula_template": "=IF(AND(NAV!G7>NAV!G8, {quarter}>5), MIN((NAV!G7-NAV!G8)*0.5, NAV!G7*0.3), 0)",
            },
            {
                "row": 5,
                "label": "Revenue",
                "pattern_type": "time_series",
                "description": "Initial revenue * (1 + growth rate)^(quarter-1)",
                "formula_template": "=Assumptions!$B$7*(1+Assumptions!$B$3)^({quarter}-1)",
            },
            {
                "row": 6,
                "label": "Management Fees",
                "pattern_type": "cross_sheet",
                "description": "Committed capital * fee rate / 4 (quarterly)",
                "formula_template": "=Assumptions!$B$8*Assumptions!$B$4/4",
            },
            {
                "row": 7,
                "label": "Net Operating Income",
                "pattern_type": "arithmetic",
                "description": "Revenue - Management Fees",
                "formula_template": "={col}5-{col}6",
            },
            {
                "row": 8,
                "label": "Cumulative Distributions",
                "pattern_type": "cumulative",
                "description": "Running total of distributions",
                "formula_template": "={prior_col}8+{col}4",
            },
            {
                "row": 9,
                "label": "Cumulative Capital Calls",
                "pattern_type": "cumulative",
                "description": "Running total of capital calls",
                "formula_template": "={prior_col}9+{col}3",
            },
        ],
    },
    "NAV": {
        "type": "calculation_sheet",
        "patterns": [
            {
                "row": 3,
                "label": "Beginning NAV",
                "pattern_type": "time_series",
                "description": "Prior period ending NAV (0 for Q1)",
                "formula_template": "={prior_col}7",
            },
            {
                "row": 4,
                "label": "+ Capital Calls",
                "pattern_type": "cross_sheet",
                "description": "Reference to Cash Flows capital calls",
                "formula_template": "='Cash Flows'!{col}3",
            },
            {
                "row": 5,
                "label": "+ NOI",
                "pattern_type": "cross_sheet",
                "description": "Reference to Cash Flows NOI",
                "formula_template": "='Cash Flows'!{col}7",
            },
            {
                "row": 6,
                "label": "- Distributions",
                "pattern_type": "cross_sheet",
                "description": "Reference to Cash Flows distributions",
                "formula_template": "='Cash Flows'!{col}4",
            },
            {
                "row": 7,
                "label": "Ending NAV",
                "pattern_type": "arithmetic",
                "description": "Beginning + Calls + NOI - Distributions",
                "formula_template": "={col}3+{col}4+{col}5-{col}6",
            },
            {
                "row": 8,
                "label": "Hurdle Amount",
                "pattern_type": "cross_sheet",
                "description": "Cumulative calls * (1 + hurdle rate * quarters/4)",
                "formula_template": "='Cash Flows'!{col}9*(1+Assumptions!$B$6*({quarter}/4))",
            },
            {
                "row": 9,
                "label": "Above Hurdle?",
                "pattern_type": "conditional",
                "description": "TRUE if Ending NAV > Hurdle Amount",
                "formula_template": "=IF({col}7>{col}8, \"TRUE\", \"FALSE\")",
            },
        ],
    },
}


if __name__ == "__main__":
    # Generate file for manual inspection
    output = Path("./synthetic_pe_fund_e2e.xlsx")
    create_synthetic_pe_fund(str(output))
    print(f"Created: {output.absolute()}")
    print("\nExpected patterns for interview:")
    for sheet, info in EXPECTED_PATTERNS.items():
        print(f"\n{sheet} ({info['type']}):")
        if "patterns" in info:
            for p in info["patterns"]:
                print(f"  Row {p['row']}: {p['label']} - {p['pattern_type']}")
                print(f"    {p['description']}")
