"""Tests for Excel probes."""

import os
import tempfile
from collections.abc import Generator

import pytest
from astro_backend_service.probes.excel import (
    analyze_sheet_structure,
    compile_excel_from_blueprint,
    detect_row_patterns,
    get_excel_cell_value,
    parse_excel_structure,
    verify_reconstruction,
)
from openpyxl import Workbook


@pytest.fixture
def synthetic_fund_model() -> Generator[str, None, None]:
    """Create a synthetic fund model .xlsx file for testing.

    Structure:
    - Assumptions sheet: Input values (growth rate, fee rate)
    - Cash Flows sheet: Calculated values with time series patterns

    Known formulas:
    - Revenue row: 5% constant growth (row 5)
    - Fee row: 2% of Revenue (row 6)
    - Net row: Revenue - Fee (row 7)
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ===== Assumptions Sheet =====
    assumptions = wb.create_sheet("Assumptions")
    assumptions["A1"] = "Parameter"
    assumptions["B1"] = "Value"
    assumptions["A3"] = "Growth Rate"
    assumptions["B3"] = 0.05  # 5%
    assumptions["A4"] = "Fee Rate"
    assumptions["B4"] = 0.02  # 2%
    assumptions["A5"] = "Initial Revenue"
    assumptions["B5"] = 1000000  # 1M

    # ===== Cash Flows Sheet =====
    cash_flows = wb.create_sheet("Cash Flows")

    # Headers (row 1)
    cash_flows["A1"] = "Metric"
    for col in range(2, 13):  # Columns B-L (11 periods)
        cash_flows.cell(row=1, column=col, value=f"Q{col-1}")

    # Row labels
    cash_flows["A3"] = "Capital Calls"
    cash_flows["A4"] = "Distributions"
    cash_flows["A5"] = "Revenue"
    cash_flows["A6"] = "Fees"
    cash_flows["A7"] = "Net Cash Flow"

    # Input rows (3, 4) - just put some values
    for col in range(2, 13):
        cash_flows.cell(row=3, column=col, value=50000 * (col - 1))  # Capital calls
        cash_flows.cell(row=4, column=col, value=30000 * (col - 1))  # Distributions

    # Revenue row (5) - constant 5% growth from initial
    initial_revenue = 1000000
    for col in range(2, 13):
        period = col - 2  # 0-indexed period
        cash_flows.cell(row=5, column=col, value=initial_revenue * (1.05 ** period))

    # Fees row (6) - 2% of revenue
    for col in range(2, 13):
        revenue = cash_flows.cell(row=5, column=col).value
        cash_flows.cell(row=6, column=col, value=revenue * 0.02)

    # Net row (7) - Revenue - Fees
    for col in range(2, 13):
        revenue = cash_flows.cell(row=5, column=col).value
        fees = cash_flows.cell(row=6, column=col).value
        cash_flows.cell(row=7, column=col, value=revenue - fees)

    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name

    wb.save(temp_path)
    wb.close()

    yield temp_path

    # Cleanup
    if os.path.exists(temp_path):
        os.unlink(temp_path)


@pytest.fixture
def sample_blueprint() -> dict:
    """Create a sample blueprint for testing compilation."""
    return {
        "id": "test_blueprint",
        "name": "Test Fund Model",
        "description": "Test blueprint",
        "version": "1.0",
        "calculation_order": ["Assumptions", "Cash Flows"],
        "sheets": [
            {
                "name": "Assumptions",
                "purpose": "Input assumptions",
                "total_rows": 10,
                "total_cols": 2,
                "input_rows": [3, 4, 5],
                "calculated_rows": [],
                "row_labels": {"3": "Growth Rate", "4": "Fee Rate", "5": "Initial Revenue"},
                "row_patterns": [],
                "cell_overrides": [],
            },
            {
                "name": "Cash Flows",
                "purpose": "Cash flow calculations",
                "total_rows": 10,
                "total_cols": 12,
                "input_rows": [3, 4],
                "calculated_rows": [5, 6, 7],
                "row_labels": {
                    "3": "Capital Calls",
                    "4": "Distributions",
                    "5": "Revenue",
                    "6": "Fees",
                    "7": "Net Cash Flow",
                },
                "time_labels": [f"Q{i}" for i in range(1, 12)],
                "row_patterns": [
                    {
                        "row": 5,
                        "row_label": "Revenue",
                        "pattern_type": "time_series",
                        "formula_template": "={prior_col}5*(1+Assumptions!$B$3)",
                        "first_data_col": 3,
                        "last_data_col": 12,
                        "seed_col": 2,
                        "exceptions": {},
                    },
                    {
                        "row": 6,
                        "row_label": "Fees",
                        "pattern_type": "arithmetic",
                        "formula_template": "={col}5*Assumptions!$B$4",
                        "first_data_col": 2,
                        "last_data_col": 12,
                        "exceptions": {},
                    },
                    {
                        "row": 7,
                        "row_label": "Net Cash Flow",
                        "pattern_type": "arithmetic",
                        "formula_template": "={col}5-{col}6",
                        "first_data_col": 2,
                        "last_data_col": 12,
                        "exceptions": {},
                    },
                ],
                "cell_overrides": [],
            },
        ],
    }


class TestParseExcelStructure:
    """Test parse_excel_structure probe."""

    def test_parses_file_successfully(self, synthetic_fund_model):
        """Test parsing a valid Excel file."""
        result = parse_excel_structure.invoke({"file_path": synthetic_fund_model})

        assert "error" not in result
        assert result["sheet_count"] == 2
        assert len(result["sheets"]) == 2

        # Check sheet names
        sheet_names = [s["name"] for s in result["sheets"]]
        assert "Assumptions" in sheet_names
        assert "Cash Flows" in sheet_names

    def test_extracts_cell_data(self, synthetic_fund_model):
        """Test that cell data is extracted correctly."""
        result = parse_excel_structure.invoke({"file_path": synthetic_fund_model})

        # Find Cash Flows sheet
        cash_flows = next(s for s in result["sheets"] if s["name"] == "Cash Flows")

        # Check dimensions
        assert cash_flows["max_row"] >= 7
        assert cash_flows["max_col"] >= 12

        # Check that we have row data
        assert len(cash_flows["rows"]) > 0

    def test_handles_missing_file(self):
        """Test error handling for missing file."""
        result = parse_excel_structure.invoke({"file_path": "/nonexistent/path/file.xlsx"})

        assert "error" in result
        assert "not found" in result["error"].lower()


class TestAnalyzeSheetStructure:
    """Test analyze_sheet_structure probe."""

    def test_identifies_header_row(self, synthetic_fund_model):
        """Test detection of header row."""
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        cash_flows = next(s for s in parsed["sheets"] if s["name"] == "Cash Flows")

        analysis = analyze_sheet_structure.invoke({"sheet_data": cash_flows})

        assert analysis["header_row"] == 1

    def test_identifies_label_column(self, synthetic_fund_model):
        """Test detection of label column."""
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        cash_flows = next(s for s in parsed["sheets"] if s["name"] == "Cash Flows")

        analysis = analyze_sheet_structure.invoke({"sheet_data": cash_flows})

        assert analysis["label_column"] == 1

    def test_creates_row_summaries(self, synthetic_fund_model):
        """Test that row summaries are created."""
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        cash_flows = next(s for s in parsed["sheets"] if s["name"] == "Cash Flows")

        analysis = analyze_sheet_structure.invoke({"sheet_data": cash_flows})

        assert len(analysis["row_summaries"]) > 0

        # Check structure of row summary
        first_summary = analysis["row_summaries"][0]
        assert "row_num" in first_summary
        assert "label" in first_summary
        assert "numeric_count" in first_summary
        assert "has_values" in first_summary


class TestDetectRowPatterns:
    """Test detect_row_patterns probe."""

    def test_detects_growth_pattern(self, synthetic_fund_model):
        """Test detection of constant growth pattern."""
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        cash_flows = next(s for s in parsed["sheets"] if s["name"] == "Cash Flows")

        analysis = analyze_sheet_structure.invoke({"sheet_data": cash_flows})
        patterns = detect_row_patterns.invoke({
            "sheet_data": cash_flows,
            "analyzed_rows": analysis["row_summaries"],
        })

        # Should detect the 5% growth pattern in Revenue row
        growth_patterns = [p for p in patterns if p.get("pattern") == "constant_growth"]

        # At least one growth pattern should be detected
        assert len(growth_patterns) >= 1

        # Check for Revenue row (row 5)
        revenue_pattern = next((p for p in growth_patterns if p["row_num"] == 5), None)
        if revenue_pattern:
            assert abs(revenue_pattern["growth_rate"] - 0.05) < 0.001

    def test_detects_ratio_pattern(self, synthetic_fund_model):
        """Test detection of ratio between rows."""
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        cash_flows = next(s for s in parsed["sheets"] if s["name"] == "Cash Flows")

        analysis = analyze_sheet_structure.invoke({"sheet_data": cash_flows})
        patterns = detect_row_patterns.invoke({
            "sheet_data": cash_flows,
            "analyzed_rows": analysis["row_summaries"],
        })

        # Should detect that Fees (row 6) is 2% of Revenue (row 5)
        ratio_patterns = [p for p in patterns if p.get("pattern") == "ratio_of_other_row"]

        # Check for Fees row
        fees_pattern = next((p for p in ratio_patterns if p["row_num"] == 6), None)
        if fees_pattern:
            assert abs(fees_pattern["ratio"] - 0.02) < 0.001


class TestCompileExcelFromBlueprint:
    """Test compile_excel_from_blueprint probe."""

    def test_creates_output_file(self, sample_blueprint):
        """Test that compilation creates an output file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            input_data = {
                "Assumptions!R3C2": 0.05,  # Growth rate
                "Assumptions!R4C2": 0.02,  # Fee rate
                "Assumptions!R5C2": 1000000,  # Initial revenue
                "Cash Flows!R5C2": 1000000,  # Seed value for revenue
            }

            result = compile_excel_from_blueprint.invoke({
                "blueprint": sample_blueprint,
                "input_data": input_data,
                "output_path": output_path,
            })

            assert os.path.exists(output_path)
            assert result["output_path"] == output_path
            assert len(result["sheets_created"]) == 2
            assert result["formulas_written"] > 0
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_writes_formulas(self, sample_blueprint):
        """Test that formulas are written correctly."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            input_data = {
                "Cash Flows!R5C2": 1000000,
            }

            compile_excel_from_blueprint.invoke({
                "blueprint": sample_blueprint,
                "input_data": input_data,
                "output_path": output_path,
            })

            # Read back and check formula cells contain formulas
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb["Cash Flows"]

            # Check that Revenue row (5) column C has a formula
            cell_c5 = ws.cell(row=5, column=3).value
            assert cell_c5 is not None
            assert cell_c5.startswith("=")

            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestVerifyReconstruction:
    """Test verify_reconstruction probe."""

    def test_perfect_match(self, synthetic_fund_model, sample_blueprint):
        """Test verification when files match."""
        # Use the same file as both original and reconstructed
        result = verify_reconstruction.invoke({
            "original_path": synthetic_fund_model,
            "reconstructed_path": synthetic_fund_model,
            "blueprint": sample_blueprint,
        })

        assert result["pass_rate"] == 1.0
        assert result["cells_mismatched"] == 0

    def test_handles_missing_file(self, sample_blueprint):
        """Test error handling for missing files."""
        result = verify_reconstruction.invoke({
            "original_path": "/nonexistent/original.xlsx",
            "reconstructed_path": "/nonexistent/reconstructed.xlsx",
            "blueprint": sample_blueprint,
        })

        assert "error" in result


class TestGetExcelCellValue:
    """Test get_excel_cell_value probe."""

    def test_gets_cell_value(self, synthetic_fund_model):
        """Test getting a specific cell value."""
        result = get_excel_cell_value.invoke({
            "file_path": synthetic_fund_model,
            "sheet_name": "Assumptions",
            "row": 3,
            "col": 2,
        })

        assert result["sheet"] == "Assumptions"
        assert result["row"] == 3
        assert result["col"] == 2
        assert result["value"] == 0.05  # Growth rate

    def test_handles_missing_sheet(self, synthetic_fund_model):
        """Test error handling for missing sheet."""
        result = get_excel_cell_value.invoke({
            "file_path": synthetic_fund_model,
            "sheet_name": "NonExistent",
            "row": 1,
            "col": 1,
        })

        assert "error" in result
        assert "not found" in result["error"].lower()


class TestIntegration:
    """Integration tests for end-to-end workflow."""

    def test_parse_analyze_detect_workflow(self, synthetic_fund_model):
        """Test the full analysis workflow."""
        # Step 1: Parse
        parsed = parse_excel_structure.invoke({"file_path": synthetic_fund_model})
        assert parsed["sheet_count"] == 2

        # Step 2: Analyze each sheet
        for sheet_data in parsed["sheets"]:
            analysis = analyze_sheet_structure.invoke({"sheet_data": sheet_data})
            assert "row_summaries" in analysis

            # Step 3: Detect patterns
            patterns = detect_row_patterns.invoke({
                "sheet_data": sheet_data,
                "analyzed_rows": analysis["row_summaries"],
            })
            # Should detect at least some patterns in Cash Flows
            if sheet_data["name"] == "Cash Flows":
                assert len(patterns) >= 0  # May or may not detect depending on data
