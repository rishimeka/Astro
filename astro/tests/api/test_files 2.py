"""Tests for file upload routes."""

import os
import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from astro_backend_service.api.main import create_app
from astro_backend_service.api.routes import files


@pytest.fixture
def client():
    """Create a test client."""
    app = create_app()
    return TestClient(app)


@pytest.fixture
def temp_upload_dir(tmp_path, monkeypatch):
    """Use a temporary directory for uploads."""
    monkeypatch.setattr(files, "UPLOAD_DIR", tmp_path)
    return tmp_path


@pytest.fixture
def sample_xlsx(tmp_path_factory):
    """Create a sample .xlsx file for testing."""
    from openpyxl import Workbook

    # Use a separate temp dir for sample files
    sample_dir = tmp_path_factory.mktemp("samples")

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    ws["B1"] = 123

    file_path = sample_dir / "test.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def sample_csv(tmp_path_factory):
    """Create a sample .csv file for testing."""
    # Use a separate temp dir for sample files
    sample_dir = tmp_path_factory.mktemp("samples")
    file_path = sample_dir / "test.csv"
    file_path.write_text("name,value\ntest,123")
    return file_path


class TestUploadFiles:
    """Test POST /files endpoint."""

    def test_upload_single_xlsx(self, client, temp_upload_dir, sample_xlsx):
        """Test uploading a single Excel file."""
        with open(sample_xlsx, "rb") as f:
            response = client.post(
                "/files",
                files={"files": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert response.status_code == 200
        data = response.json()

        assert data["message"] == "Successfully uploaded 1 file(s)"
        assert len(data["files"]) == 1

        uploaded = data["files"][0]
        assert uploaded["original_name"] == "test.xlsx"
        assert uploaded["mime_type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert uploaded["size_bytes"] > 0
        assert Path(uploaded["stored_path"]).exists()

    def test_upload_multiple_files(self, client, temp_upload_dir, sample_xlsx, sample_csv):
        """Test uploading multiple files at once."""
        with open(sample_xlsx, "rb") as f1, open(sample_csv, "rb") as f2:
            response = client.post(
                "/files",
                files=[
                    ("files", ("test.xlsx", f1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                    ("files", ("test.csv", f2, "text/csv")),
                ],
            )

        assert response.status_code == 200
        data = response.json()

        assert data["message"] == "Successfully uploaded 2 file(s)"
        assert len(data["files"]) == 2

    def test_upload_disallowed_extension(self, client, temp_upload_dir, tmp_path):
        """Test that disallowed file types are rejected."""
        bad_file = tmp_path / "test.exe"
        bad_file.write_bytes(b"not an executable")

        with open(bad_file, "rb") as f:
            response = client.post(
                "/files",
                files={"files": ("test.exe", f, "application/octet-stream")},
            )

        assert response.status_code == 400
        assert "not allowed" in response.json()["detail"]

    def test_upload_stores_file_correctly(self, client, temp_upload_dir, sample_xlsx):
        """Test that file content is preserved on upload."""
        original_content = sample_xlsx.read_bytes()

        with open(sample_xlsx, "rb") as f:
            response = client.post(
                "/files",
                files={"files": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert response.status_code == 200
        stored_path = Path(response.json()["files"][0]["stored_path"])
        stored_content = stored_path.read_bytes()

        assert original_content == stored_content


class TestGetFileMetadata:
    """Test GET /files/{file_id} endpoint."""

    def test_get_existing_file(self, client, temp_upload_dir, sample_xlsx):
        """Test retrieving metadata for an uploaded file."""
        # Upload first
        with open(sample_xlsx, "rb") as f:
            upload_response = client.post(
                "/files",
                files={"files": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        file_id = upload_response.json()["files"][0]["id"]

        # Get metadata
        response = client.get(f"/files/{file_id}")

        assert response.status_code == 200
        data = response.json()
        assert data["id"] == file_id

    def test_get_nonexistent_file(self, client, temp_upload_dir):
        """Test retrieving a file that doesn't exist."""
        response = client.get("/files/nonexistent-id")

        assert response.status_code == 404
        assert "not found" in response.json()["detail"]


class TestDeleteFile:
    """Test DELETE /files/{file_id} endpoint."""

    def test_delete_existing_file(self, client, temp_upload_dir, sample_xlsx):
        """Test deleting an uploaded file."""
        # Upload first
        with open(sample_xlsx, "rb") as f:
            upload_response = client.post(
                "/files",
                files={"files": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        file_id = upload_response.json()["files"][0]["id"]
        stored_path = Path(upload_response.json()["files"][0]["stored_path"])

        assert stored_path.exists()

        # Delete
        response = client.delete(f"/files/{file_id}")

        assert response.status_code == 200
        assert not stored_path.exists()

    def test_delete_nonexistent_file(self, client, temp_upload_dir):
        """Test deleting a file that doesn't exist."""
        response = client.delete("/files/nonexistent-id")

        assert response.status_code == 404


class TestListFiles:
    """Test GET /files endpoint."""

    def test_list_empty(self, client, temp_upload_dir):
        """Test listing files when none exist."""
        response = client.get("/files")

        assert response.status_code == 200
        data = response.json()
        assert data["count"] == 0
        assert data["files"] == []

    def test_list_uploaded_files(self, client, temp_upload_dir, sample_xlsx, sample_csv):
        """Test listing all uploaded files."""
        # Upload files
        with open(sample_xlsx, "rb") as f:
            client.post(
                "/files",
                files={"files": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        with open(sample_csv, "rb") as f:
            client.post(
                "/files",
                files={"files": ("test.csv", f, "text/csv")},
            )

        # List
        response = client.get("/files")

        assert response.status_code == 200
        data = response.json()
        assert data["count"] == 2
        assert len(data["files"]) == 2


class TestIntegration:
    """Integration tests for file upload flow."""

    def test_upload_get_path_for_constellation(self, client, temp_upload_dir, sample_xlsx):
        """Test the full flow: upload file, get path for constellation run."""
        # 1. Upload file
        with open(sample_xlsx, "rb") as f:
            upload_response = client.post(
                "/files",
                files={"files": ("fund_model.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert upload_response.status_code == 200
        file_metadata = upload_response.json()["files"][0]

        # 2. Get the stored path (this would be passed to constellation run)
        stored_path = file_metadata["stored_path"]
        assert Path(stored_path).exists()
        assert stored_path.endswith(".xlsx")

        # 3. Verify the path works with openpyxl (as Excel probes would use)
        from openpyxl import load_workbook

        wb = load_workbook(stored_path)
        assert wb.active["A1"].value == "Test"
        wb.close()
