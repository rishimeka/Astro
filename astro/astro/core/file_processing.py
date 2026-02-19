"""
File processing utilities for constellation runs and launchpad.

Handles:
- Excel files (.xlsx, .xls) → Parsed JSON structure
- PDF files (.pdf) → Binary data for model native support
- Other files → Basic metadata
"""

import json
import logging
from enum import Enum
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class FileType(str, Enum):
    """Supported file types."""

    EXCEL = "excel"
    PDF = "pdf"
    IMAGE = "image"
    TEXT = "text"
    UNKNOWN = "unknown"


class ProcessedFile:
    """Container for processed file data."""

    def __init__(
        self,
        file_type: FileType,
        original_filename: str,
        file_path: str,
        data: Any = None,
        metadata: dict[str, Any] | None = None,
    ):
        self.file_type = file_type
        self.original_filename = original_filename
        self.file_path = file_path
        self.data = data
        self.metadata = metadata or {}

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "file_type": self.file_type,
            "original_filename": self.original_filename,
            "file_path": self.file_path,
            "data": self.data,
            "metadata": self.metadata,
        }


def detect_file_type(filename: str) -> FileType:
    """Detect file type from filename extension."""
    suffix = Path(filename).suffix.lower()

    if suffix in [".xlsx", ".xls", ".xlsm"]:
        return FileType.EXCEL
    elif suffix == ".pdf":
        return FileType.PDF
    elif suffix in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
        return FileType.IMAGE
    elif suffix in [".txt", ".md", ".csv"]:
        return FileType.TEXT
    else:
        return FileType.UNKNOWN


def process_excel_file(file_path: str) -> dict[str, Any]:
    """
    Parse Excel file into JSON structure.

    Returns dict with sheets, rows, columns, and values.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Cannot process Excel files.")
        raise ImportError(
            "openpyxl is required for Excel processing. Install with: pip install openpyxl"
        )

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        result = {"sheets": []}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = {
                "name": sheet_name,
                "rows": [],
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }

            # Read all rows
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Convert row tuple to list, handling None values
                row_data = [
                    cell if cell is not None else "" for cell in row
                ]
                sheet_data["rows"].append({"row_num": row_idx, "values": row_data})

            result["sheets"].append(sheet_data)

        return result

    except Exception as e:
        logger.error(f"Error processing Excel file {file_path}: {e}")
        raise


def process_file(file_path: str, original_filename: str) -> ProcessedFile:
    """
    Process uploaded file based on type.

    Args:
        file_path: Path to the uploaded file
        original_filename: Original filename from upload

    Returns:
        ProcessedFile with parsed data and metadata
    """
    file_type = detect_file_type(original_filename)

    logger.info(
        f"Processing file: {original_filename} (type: {file_type}, path: {file_path})"
    )

    if file_type == FileType.EXCEL:
        # Parse Excel into JSON structure
        try:
            data = process_excel_file(file_path)
            return ProcessedFile(
                file_type=file_type,
                original_filename=original_filename,
                file_path=file_path,
                data=data,
                metadata={"sheets_count": len(data["sheets"])},
            )
        except Exception as e:
            logger.error(f"Failed to process Excel file: {e}")
            return ProcessedFile(
                file_type=file_type,
                original_filename=original_filename,
                file_path=file_path,
                data=None,
                metadata={"error": str(e)},
            )

    elif file_type == FileType.PDF:
        # For PDFs, just store the path - models will handle natively
        file_size = Path(file_path).stat().st_size
        return ProcessedFile(
            file_type=file_type,
            original_filename=original_filename,
            file_path=file_path,
            data=None,  # Will be read by model API
            metadata={"size_bytes": file_size},
        )

    elif file_type == FileType.IMAGE:
        # For images, store path - models will handle natively
        file_size = Path(file_path).stat().st_size
        return ProcessedFile(
            file_type=file_type,
            original_filename=original_filename,
            file_path=file_path,
            data=None,  # Will be read by model API
            metadata={"size_bytes": file_size},
        )

    elif file_type == FileType.TEXT:
        # Read text files directly
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            return ProcessedFile(
                file_type=file_type,
                original_filename=original_filename,
                file_path=file_path,
                data={"content": content},
                metadata={"size_bytes": len(content)},
            )
        except Exception as e:
            logger.error(f"Failed to read text file: {e}")
            return ProcessedFile(
                file_type=file_type,
                original_filename=original_filename,
                file_path=file_path,
                data=None,
                metadata={"error": str(e)},
            )

    else:
        # Unknown file type - just store metadata
        file_size = Path(file_path).stat().st_size
        return ProcessedFile(
            file_type=file_type,
            original_filename=original_filename,
            file_path=file_path,
            data=None,
            metadata={"size_bytes": file_size, "warning": "Unknown file type"},
        )


def format_file_context_for_llm(processed_file: ProcessedFile) -> str:
    """
    Format processed file data as context string for LLM.

    Args:
        processed_file: The processed file data

    Returns:
        Formatted string suitable for LLM context
    """
    if processed_file.file_type == FileType.EXCEL and processed_file.data:
        # Format Excel data as readable structure
        context = f"Uploaded Excel File: {processed_file.original_filename}\n\n"

        for sheet in processed_file.data.get("sheets", []):
            context += f"Sheet: {sheet['name']}\n"
            context += f"Dimensions: {sheet['max_row']} rows × {sheet['max_column']} columns\n\n"

            # Include first few rows as preview
            preview_rows = sheet["rows"][:10]
            for row_data in preview_rows:
                row_num = row_data["row_num"]
                values = row_data["values"]
                context += f"Row {row_num}: {values}\n"

            if len(sheet["rows"]) > 10:
                context += f"... ({len(sheet['rows']) - 10} more rows)\n"

            context += "\n"

        return context

    elif processed_file.file_type == FileType.TEXT and processed_file.data:
        # Include text content directly
        content = processed_file.data.get("content", "")
        return (
            f"Uploaded Text File: {processed_file.original_filename}\n\n{content}\n"
        )

    elif processed_file.file_type in [FileType.PDF, FileType.IMAGE]:
        # For PDF/images, just note availability (will be passed to model directly)
        return f"Uploaded {processed_file.file_type.value.upper()} file: {processed_file.original_filename} (available for analysis)"

    else:
        return f"Uploaded file: {processed_file.original_filename} (type: {processed_file.file_type.value})"
