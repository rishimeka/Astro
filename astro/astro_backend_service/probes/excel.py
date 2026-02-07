"""Excel probes for fund model reverse-engineering.

Provides probes for:
- Parsing Excel file structure
- Analyzing sheet structure and detecting patterns
- Compiling blueprints to Excel with formulas
- Verifying reconstruction accuracy
"""

import os
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from astro_backend_service.probes.decorator import probe


@probe
def parse_excel_structure(file_path: str) -> Dict[str, Any]:
    """Parse an Excel file and extract its complete structure.

    Reads an Excel file and extracts sheet names, cell values, data types,
    and layout information. Returns a structured representation suitable
    for AI analysis.

    Args:
        file_path: Absolute path to the .xlsx file to parse.

    Returns:
        Dictionary containing:
        - file_name: Name of the file
        - sheet_count: Number of sheets
        - sheets: List of sheet data with rows, dimensions, etc.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}

    # Load with data_only=True to get computed values (not formulas)
    wb = load_workbook(file_path, data_only=True)

    result = {
        "file_name": os.path.basename(file_path),
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": [],
        }

        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False
        ):
            row_data = []
            for cell in row:
                cell_info = {
                    "row": cell.row,
                    "col": cell.column,
                    "col_letter": cell.column_letter,
                    "value": cell.value,
                    "data_type": cell.data_type,  # 'n'=number, 's'=string, 'd'=date
                    "number_format": cell.number_format,
                    "is_merged": isinstance(cell, MergedCell),
                }
                row_data.append(cell_info)
            sheet_data["rows"].append(row_data)

        result["sheets"].append(sheet_data)

    wb.close()
    return result


@probe
def analyze_sheet_structure(sheet_data: Dict[str, Any]) -> Dict[str, Any]:
    """Analyze a single sheet's structure to identify patterns.

    Examines a sheet's data to identify headers, labels, data regions,
    numeric patterns, and potential formula relationships.

    Args:
        sheet_data: Sheet data from parse_excel_structure (single sheet).

    Returns:
        Dictionary containing:
        - sheet_name: Name of the sheet
        - header_row: Detected header row number
        - label_column: Detected label column number
        - data_region: Boundaries of the data region
        - row_summaries: Analysis of each row
        - potential_patterns: Detected numeric patterns
    """
    rows = sheet_data.get("rows", [])
    name = sheet_data.get("name", "Unknown")

    analysis = {
        "sheet_name": name,
        "header_row": None,
        "label_column": None,
        "data_region": None,
        "row_summaries": [],
        "potential_patterns": [],
        "numeric_relationships": [],
    }

    if not rows:
        return analysis

    # Identify header row (first row that's mostly strings)
    for i, row in enumerate(rows):
        values = [c["value"] for c in row if c["value"] is not None]
        if all(isinstance(v, str) for v in values) and len(values) > 2:
            analysis["header_row"] = i + 1
            break

    # Identify label column (first column that's mostly strings)
    col_0_values = [row[0]["value"] for row in rows if row and row[0]["value"] is not None]
    if col_0_values and sum(isinstance(v, str) for v in col_0_values) > len(col_0_values) * 0.7:
        analysis["label_column"] = 1

    # Summarize each row
    for i, row in enumerate(rows):
        numeric_vals = [
            c["value"] for c in row if isinstance(c["value"], (int, float))
        ]
        label = row[0]["value"] if row and row[0]["value"] else f"Row {i + 1}"

        row_summary = {
            "row_num": i + 1,
            "label": str(label),
            "numeric_count": len(numeric_vals),
            "has_values": len(numeric_vals) > 0,
            "sample_values": numeric_vals[:5] if numeric_vals else [],
            "all_zero": all(v == 0 for v in numeric_vals) if numeric_vals else False,
            "is_likely_total": _is_likely_total(label, numeric_vals),
        }
        analysis["row_summaries"].append(row_summary)

    # Detect column-wise patterns (time series)
    for i, row in enumerate(rows):
        numeric_vals = [
            c["value"] for c in row if isinstance(c["value"], (int, float))
        ]
        if len(numeric_vals) >= 3:
            pattern = _detect_numeric_pattern(numeric_vals)
            if pattern:
                analysis["potential_patterns"].append({
                    "row_num": i + 1,
                    "label": str(row[0]["value"]) if row and row[0]["value"] else f"Row {i + 1}",
                    "pattern": pattern,
                })

    return analysis


def _detect_numeric_pattern(values: List[float]) -> Optional[Dict[str, Any]]:
    """Detect if values follow a recognizable pattern."""
    if len(values) < 3:
        return None

    # Filter out zero/None values for ratio calculation
    non_zero_vals = [v for v in values if v and v != 0]
    if len(non_zero_vals) < 3:
        return None

    # Check constant growth (constant ratio between consecutive values)
    ratios = []
    for i in range(len(non_zero_vals) - 1):
        if non_zero_vals[i] != 0:
            ratios.append(non_zero_vals[i + 1] / non_zero_vals[i])

    if ratios and len(ratios) >= 2:
        # Check if all ratios are approximately equal
        avg_ratio = sum(ratios) / len(ratios)
        if all(abs(r - avg_ratio) < 0.001 for r in ratios):
            return {"type": "constant_growth", "growth_rate": avg_ratio - 1}

    # Check constant value
    if all(abs(v - values[0]) < 0.01 for v in values if v is not None):
        return {"type": "constant", "value": values[0]}

    return None


def _is_likely_total(label: Any, values: List[float]) -> bool:
    """Heuristic: is this row a subtotal or total line?"""
    label_str = str(label).lower() if label else ""
    total_keywords = ["total", "subtotal", "sum", "net", "gross"]
    return any(kw in label_str for kw in total_keywords)


@probe
def detect_row_patterns(
    sheet_data: Dict[str, Any],
    analyzed_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Analyze numeric data to detect repeating patterns across rows.

    Identifies rows that likely share the same formula logic, reducing
    the number of questions the interviewer needs to ask.

    Args:
        sheet_data: Raw sheet data from parse_excel_structure.
        analyzed_rows: Row summaries from analyze_sheet_structure.

    Returns:
        List of detected patterns with suggested formulas and confidence.
    """
    patterns = []
    rows = sheet_data.get("rows", [])

    for row_info in analyzed_rows:
        if not row_info.get("has_values"):
            continue

        row_idx = row_info["row_num"] - 1
        if row_idx >= len(rows):
            continue

        row = rows[row_idx]

        numeric_cells = [
            (c["col"], c["value"])
            for c in row
            if isinstance(c["value"], (int, float)) and c["value"] != 0
        ]

        if len(numeric_cells) < 3:
            continue

        vals = [v for _, v in numeric_cells]

        # Test: constant ratio to prior column (growth pattern)
        ratios = []
        for i in range(1, len(vals)):
            if vals[i - 1] != 0:
                ratios.append(vals[i] / vals[i - 1])

        if ratios and len(set(round(r, 4) for r in ratios)) == 1:
            growth_rate = round(ratios[0] - 1, 6)
            patterns.append({
                "row_num": row_info["row_num"],
                "label": row_info["label"],
                "pattern": "constant_growth",
                "growth_rate": growth_rate,
                "confidence": 0.8,
                "suggested_formula": "={prior_col}{row}*(1+GROWTH_RATE)",
                "needs_confirmation": True,
                "question": (
                    f"Row '{row_info['label']}' shows a constant "
                    f"{round(growth_rate * 100, 2)}% growth rate each period. "
                    f"Is this correct? What drives this growth rate?"
                ),
            })
            continue

        # Test: ratio to another row (e.g., margin calculation)
        for other_info in analyzed_rows:
            if other_info["row_num"] == row_info["row_num"]:
                continue
            if not other_info.get("has_values"):
                continue

            other_idx = other_info["row_num"] - 1
            if other_idx >= len(rows):
                continue

            other_row = rows[other_idx]
            other_vals = [
                c["value"]
                for c in other_row
                if isinstance(c["value"], (int, float))
            ]

            if len(other_vals) != len(vals):
                continue

            cross_ratios = []
            for v1, v2 in zip(vals, other_vals):
                if v2 != 0:
                    cross_ratios.append(v1 / v2)

            if cross_ratios and len(set(round(r, 4) for r in cross_ratios)) == 1:
                ratio = round(cross_ratios[0], 6)
                patterns.append({
                    "row_num": row_info["row_num"],
                    "label": row_info["label"],
                    "pattern": "ratio_of_other_row",
                    "other_row": other_info["row_num"],
                    "other_label": other_info["label"],
                    "ratio": ratio,
                    "confidence": 0.7,
                    "suggested_formula": f"={other_info['label'].replace(' ', '_')}*{round(ratio, 4)}",
                    "needs_confirmation": True,
                    "question": (
                        f"Row '{row_info['label']}' appears to be "
                        f"{round(ratio * 100, 2)}% of '{other_info['label']}' "
                        f"in every period. Is this a fixed ratio? "
                        f"Where does the {round(ratio * 100, 2)}% come from?"
                    ),
                })
                break

    return patterns


@probe
def compile_excel_from_blueprint(
    blueprint: Dict[str, Any],
    input_data: Dict[str, Any],
    output_path: str,
) -> Dict[str, Any]:
    """Compile a ModelBlueprint and input data into a working .xlsx file.

    Takes a blueprint (with formula templates) and input data, produces
    a working Excel file with actual formulas. This is deterministic -
    no LLM calls.

    Args:
        blueprint: ModelBlueprint dict with sheets, patterns, and relationships.
        input_data: Dictionary of input values keyed by "Sheet!R{row}C{col}".
        output_path: Where to save the compiled .xlsx file.

    Returns:
        Dictionary containing:
        - output_path: Path to the created file
        - sheets_created: List of sheet names
        - formulas_written: Count of formula cells
        - values_written: Count of value cells
        - errors: List of any errors encountered
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    report = {
        "sheets_created": [],
        "formulas_written": 0,
        "values_written": 0,
        "errors": [],
    }

    calculation_order = blueprint.get("calculation_order", [])
    sheets = blueprint.get("sheets", [])

    # Create sheets in calculation order
    for sheet_name in calculation_order:
        sheet_spec = next(
            (s for s in sheets if s["name"] == sheet_name), None
        )
        if not sheet_spec:
            report["errors"].append(f"Sheet spec not found: {sheet_name}")
            continue

        ws = wb.create_sheet(title=sheet_name)

        # Write headers (time labels)
        time_labels = sheet_spec.get("time_labels", [])
        for col_idx, label in enumerate(time_labels, start=2):
            ws.cell(row=1, column=col_idx, value=label)

        # Write row labels
        row_labels = sheet_spec.get("row_labels", {})
        for row_num_str, label in row_labels.items():
            ws.cell(row=int(row_num_str), column=1, value=label)

        # Apply row patterns (the bulk of the work)
        for pattern in sheet_spec.get("row_patterns", []):
            row = pattern["row"]
            template = pattern["formula_template"]

            # Write seed column as input value
            if pattern.get("seed_col"):
                seed_key = f"{sheet_name}!R{row}C{pattern['seed_col']}"
                seed_val = input_data.get(seed_key)
                if seed_val is not None:
                    ws.cell(row=row, column=pattern["seed_col"], value=seed_val)
                    report["values_written"] += 1

            # Write formula for each data column
            first_col = pattern.get("first_data_col", 2)
            last_col = pattern.get("last_data_col", sheet_spec.get("total_cols", 50))
            exceptions = pattern.get("exceptions", {})

            for col in range(first_col, last_col + 1):
                # Check for column-specific exception
                if col in exceptions:
                    formula = _resolve_formula(exceptions[col], row, col, sheet_name)
                elif str(col) in exceptions:  # Handle string keys from JSON
                    formula = _resolve_formula(exceptions[str(col)], row, col, sheet_name)
                else:
                    formula = _resolve_formula(template, row, col, sheet_name)

                ws.cell(row=row, column=col, value=formula)
                report["formulas_written"] += 1

        # Apply cell overrides
        for cell_rel in sheet_spec.get("cell_overrides", []):
            if cell_rel.get("formula_template"):
                formula = _resolve_formula(
                    cell_rel["formula_template"],
                    cell_rel["row"],
                    cell_rel["col"],
                    sheet_name,
                )
                ws.cell(row=cell_rel["row"], column=cell_rel["col"], value=formula)
                report["formulas_written"] += 1

        # Write input values
        for input_row in sheet_spec.get("input_rows", []):
            total_cols = sheet_spec.get("total_cols", 50)
            for col in range(2, total_cols + 1):
                key = f"{sheet_name}!R{input_row}C{col}"
                val = input_data.get(key)
                if val is not None:
                    ws.cell(row=input_row, column=col, value=val)
                    report["values_written"] += 1

        report["sheets_created"].append(sheet_name)

    # Save
    wb.save(output_path)
    report["output_path"] = output_path
    return report


def _resolve_formula(template: str, row: int, col: int, sheet: str) -> str:
    """Convert a formula template into an actual Excel formula.

    Replaces placeholders: {col}, {prior_col}, {row}, {col_letter}, etc.
    """
    col_letter = get_column_letter(col)
    prior_col_letter = get_column_letter(col - 1) if col > 1 else col_letter

    formula = template
    formula = formula.replace("{col}", col_letter)
    formula = formula.replace("{col_letter}", col_letter)
    formula = formula.replace("{prior_col}", prior_col_letter)
    formula = formula.replace("{col_num}", str(col))
    formula = formula.replace("{row}", str(row))
    formula = formula.replace("{prior_col_num}", str(col - 1))

    # Ensure it starts with =
    if not formula.startswith("="):
        formula = "=" + formula

    return formula


@probe
def verify_reconstruction(
    original_path: str,
    reconstructed_path: str,
    blueprint: Dict[str, Any],
) -> Dict[str, Any]:
    """Compare a reconstructed model against the original output.

    Diffs every calculated cell, reports discrepancies with context.

    Args:
        original_path: Path to the original .xlsx file (values only).
        reconstructed_path: Path to the reconstructed .xlsx file.
        blueprint: ModelBlueprint dict for knowing which cells to check.

    Returns:
        Dictionary containing:
        - total_cells_checked: Number of cells compared
        - cells_matching: Number matching within tolerance
        - cells_mismatched: Number with discrepancies
        - cells_missing: Number missing from reconstruction
        - discrepancies: List of specific mismatches
        - pass_rate: Percentage of matching cells
    """
    if not os.path.exists(original_path):
        return {"error": f"Original file not found: {original_path}"}
    if not os.path.exists(reconstructed_path):
        return {"error": f"Reconstructed file not found: {reconstructed_path}"}

    original = load_workbook(original_path, data_only=True)
    reconstructed = load_workbook(reconstructed_path, data_only=True)

    report = {
        "total_cells_checked": 0,
        "cells_matching": 0,
        "cells_mismatched": 0,
        "cells_missing": 0,
        "discrepancies": [],
        "pass_rate": 0.0,
    }

    sheets = blueprint.get("sheets", [])

    for sheet_spec in sheets:
        sheet_name = sheet_spec["name"]

        if sheet_name not in original.sheetnames:
            report["cells_missing"] += 1
            continue

        orig_ws = original[sheet_name]

        if sheet_name not in reconstructed.sheetnames:
            report["discrepancies"].append({
                "sheet": sheet_name,
                "issue": "Sheet missing from reconstruction",
            })
            continue

        recon_ws = reconstructed[sheet_name]

        # Check every calculated row
        calculated_rows = sheet_spec.get("calculated_rows", [])
        total_cols = sheet_spec.get("total_cols", 50)
        row_labels = sheet_spec.get("row_labels", {})

        for row_num in calculated_rows:
            for col in range(2, total_cols + 1):
                orig_val = orig_ws.cell(row=row_num, column=col).value
                recon_val = recon_ws.cell(row=row_num, column=col).value

                report["total_cells_checked"] += 1

                if orig_val is None and recon_val is None:
                    report["cells_matching"] += 1
                    continue

                if orig_val is None or recon_val is None:
                    report["cells_mismatched"] += 1
                    report["discrepancies"].append({
                        "sheet": sheet_name,
                        "row": row_num,
                        "col": col,
                        "row_label": row_labels.get(str(row_num), f"Row {row_num}"),
                        "expected": orig_val,
                        "actual": recon_val,
                        "issue": "Value present in one but not other",
                    })
                    continue

                # Numeric comparison with tolerance
                if isinstance(orig_val, (int, float)) and isinstance(
                    recon_val, (int, float)
                ):
                    if orig_val == 0:
                        match = abs(recon_val) < 0.01
                    else:
                        match = abs((recon_val - orig_val) / orig_val) < 0.01

                    if match:
                        report["cells_matching"] += 1
                    else:
                        report["cells_mismatched"] += 1
                        deviation = None
                        if orig_val != 0:
                            deviation = abs((recon_val - orig_val) / orig_val) * 100
                        report["discrepancies"].append({
                            "sheet": sheet_name,
                            "row": row_num,
                            "col": col,
                            "row_label": row_labels.get(str(row_num), f"Row {row_num}"),
                            "expected": orig_val,
                            "actual": recon_val,
                            "deviation_pct": deviation,
                            "issue": "Value mismatch",
                        })
                else:
                    report["cells_matching"] += 1  # Non-numeric, assume match

    if report["total_cells_checked"] > 0:
        report["pass_rate"] = report["cells_matching"] / report["total_cells_checked"]

    original.close()
    reconstructed.close()

    return report


@probe
def get_excel_cell_value(file_path: str, sheet_name: str, row: int, col: int) -> Dict[str, Any]:
    """Get a specific cell's value from an Excel file.

    Useful for debugging and verifying specific cells.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Name of the sheet.
        row: 1-indexed row number.
        col: 1-indexed column number.

    Returns:
        Dictionary with cell value, data type, and location info.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}

    wb = load_workbook(file_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"error": f"Sheet not found: {sheet_name}"}

    ws = wb[sheet_name]
    cell = ws.cell(row=row, column=col)

    result = {
        "sheet": sheet_name,
        "row": row,
        "col": col,
        "col_letter": get_column_letter(col),
        "value": cell.value,
        "data_type": cell.data_type,
        "number_format": cell.number_format,
    }

    wb.close()
    return result
